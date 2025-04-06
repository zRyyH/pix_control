from openpyxl import load_workbook
from io import BytesIO


def extract_text_from_xlsx(
    content: bytes,
    sheet_index: int = 0,
    start_row: int = 1,
    end_row: int = None,
    ignore_columns=None,
):
    """
    Converte um arquivo XLSX em uma lista de dicionários.

    Args:
        content: Conteúdo binário do arquivo XLSX.
        sheet_index: Índice da planilha a ser lida.
        start_row: Linha inicial para começar a leitura (1 = primeira linha).
        end_row: Linha final para terminar a leitura (None = até o final).
        ignore_columns: Lista de índices ou nomes de colunas a serem ignoradas.

    Returns:
        Lista de dicionários com os dados.
    """
    if ignore_columns is None:
        ignore_columns = []

    in_memory_file = BytesIO(content)
    wb = load_workbook(in_memory_file, data_only=True)
    sheet = wb.worksheets[sheet_index]

    # Determinar a última linha se end_row não for especificado
    max_row = sheet.max_row if end_row is None else end_row

    rows = list(sheet.iter_rows(min_row=start_row, max_row=max_row, values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    # Converter nomes de colunas para índices
    ignore_indices = []
    for col in ignore_columns:
        if isinstance(col, int):
            ignore_indices.append(col)
        elif isinstance(col, str) and col in headers:
            ignore_indices.append(headers.index(col))

    # Filtrar headers para remover colunas ignoradas
    filtered_headers = [
        headers[i] for i in range(len(headers)) if i not in ignore_indices
    ]

    data = []

    for row in rows[1:]:
        row_dict = {}
        for i in range(len(headers)):
            if i not in ignore_indices:
                row_dict[headers[i]] = row[i]
        data.append(row_dict)

    return data
